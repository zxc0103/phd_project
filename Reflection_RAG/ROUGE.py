#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Notes: 
      (1)计算完成后结果报告F1。
      (2)
"""
import jieba
import openpyxl
from rouge import Rouge

file_path = "./2_自动评估/2_最终评估/data.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

correct_translations = [row[0].value for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1)]
reference_translations = [row[0].value for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2)]

rouge = Rouge()

rouge_scores = []
for correct, reference in zip(correct_translations, reference_translations):
    correct_tokens = ' '.join(jieba.cut(correct))
    reference_tokens = ' '.join(jieba.cut(reference))
    scores = rouge.get_scores(correct_tokens, reference_tokens)[0]
    rouge_scores.append(scores)

for i, scores in enumerate(rouge_scores):
    print(f'句子 {i+1} 的 ROUGE 分数:')
    print(f"ROUGE-1: F1={scores['rouge-1']['f']}, Precision={scores['rouge-1']['p']}, Recall={scores['rouge-1']['r']}")
    print(f"ROUGE-2: F1={scores['rouge-2']['f']}, Precision={scores['rouge-2']['p']}, Recall={scores['rouge-2']['r']}")
    print(f"ROUGE-L: F1={scores['rouge-l']['f']}, Precision={scores['rouge-l']['p']}, Recall={scores['rouge-l']['r']}")

result_wb = openpyxl.Workbook()
result_sheet = result_wb.active
result_sheet.title = 'ROUGE Scores'
result_sheet.append(['Correct Translation', 'Reference Translation', 'ROUGE-1 F1', 'ROUGE-1 Precision', 'ROUGE-1 Recall', 'ROUGE-2 F1', 'ROUGE-2 Precision', 'ROUGE-2 Recall', 'ROUGE-L F1', 'ROUGE-L Precision', 'ROUGE-L Recall'])

for correct, reference, scores in zip(correct_translations, reference_translations, rouge_scores):
    result_sheet.append([
        correct, 
        reference, 
        scores['rouge-1']['f'], scores['rouge-1']['p'], scores['rouge-1']['r'], 
        scores['rouge-2']['f'], scores['rouge-2']['p'], scores['rouge-2']['r'], 
        scores['rouge-l']['f'], scores['rouge-l']['p'], scores['rouge-l']['r']
    ])

result_wb.save("./2_自动评估/2_最终评估/ROUGE_results.xlsx")