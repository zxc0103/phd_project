#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
运行前需要下载bertscore库和bert-base-chinese模型：
1. 安装Pytorch
1. pip3 install bert-score
2. 下载中文模型：
    （1）首先，在模型保存的根目录下载文件：http://61.133.217.142:20800/download/model_download.py （linux系统为：wget http://61.133.217.142:20800/download/model_download.py）
    （2）然后，在根目录启动终端，运行：python model_download.py --mirror --repo_id bert-base-chinese
    （3）最后，将本.py文件放在根目录运行。
3. num_layers=12 参数可自定义，也可直接删除使用默认配置。
4. 计算完成后结果报告F1。
"""
from bert_score import score
import openpyxl
import sys

INPUT_FILE = './简答题-评估数据.xlsx'
OUTPUT_FILE = './results_BERTScore.xlsx'
MODEL_TYPE = "bert-base-chinese"
LANG = "zh"

def read_pairs_from_excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    cands = []
    refs = []
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
        cand = row[0].value
        ref = row[1].value
        if cand is None or ref is None:
            continue
        cands.append(str(cand))
        refs.append(str(ref))
    return cands, refs

def save_scores_to_excel(ps, rs, f1s, out_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'results_BERTScore'
    sheet.append(['P', 'R', 'F1'])
    for p, r, f in zip(ps, rs, f1s):
        sheet.append([float(p), float(r), float(f)])
    wb.save(out_path)

def main():
    try:
        cands, refs = read_pairs_from_excel(INPUT_FILE)
    except FileNotFoundError:
        print(f"找不到文件: {INPUT_FILE}", file=sys.stderr)
        sys.exit(1)

    if not cands:
        print("未读取到有效句子对。", file=sys.stderr)
        sys.exit(1)

    print(f"读取到 {len(cands)} 条句子对，开始计算 BERTScore（model={MODEL_TYPE}）...")

    P, R, F1 = score(cands, refs, model_type=MODEL_TYPE, lang=LANG, num_layers=12, verbose=True) # num_layers=12 可自定义或删除

    try:
        ps = P.tolist()
        rs = R.tolist()
        f1s = F1.tolist()
    except Exception:
        ps = [float(x) for x in P]
        rs = [float(x) for x in R]
        f1s = [float(x) for x in F1]

    save_scores_to_excel(ps, rs, f1s, OUTPUT_FILE)
    avg_f1 = sum(f1s) / len(f1s)
    print(f"计算完成，结果已保存到 {OUTPUT_FILE}")
    print(f"平均 F1: {avg_f1:.6f}")

if __name__ == '__main__':
    main()