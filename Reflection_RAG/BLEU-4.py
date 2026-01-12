#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Notes: 
      (1)
      (2)
"""
import jieba
import openpyxl
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction

file_path = './简答题-评估数据.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

correct_translations = [row[0].value for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1)]
reference_translations = [row[0].value for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2)]

smoothie = SmoothingFunction().method4

bleu_scores = []
for correct, reference in zip(correct_translations, reference_translations):
    correct_tokens = list(jieba.cut(correct))
    reference_tokens = [list(jieba.cut(reference))]
    bleu_score = sentence_bleu(reference_tokens, correct_tokens, smoothing_function=smoothie)
    bleu_scores.append(bleu_score)

"""
# 打印每个句子的BLEU-4分数
print("\n\n\n")
for i, score in enumerate(bleu_scores):
    print(f'{score}')
print("\n\n\n")

"""

result_wb = openpyxl.Workbook()
result_sheet = result_wb.active
result_sheet.title = 'results_BLEU-4'
result_sheet.append(['result_BLEU-4'])

for score in bleu_scores:
    result_sheet.append([score])
result_wb.save('./results_BLEU-4.xlsx')